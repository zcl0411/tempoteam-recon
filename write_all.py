import json, os

# Write rules
rules = {
  "country": "belgium",
  "supplier": "TEMPOTEAM",
  "label": "Belgium - TEMPOTEAM",
  "rules": {
    "lunch_deduction": {"enabled": True, "description": "Deduct 0.5h lunch for shifts >= 6h", "max_daily_diff": 0.5, "per_period": True},
    "supplement_equipe": {"enabled": True, "description": "Shift allowance 7.5% for 6h-14h/14h-22h shifts", "rate_pct": 7.5, "qualifying_start_hours": [6, 14]},
    "dimona_social_security": {"enabled": True, "description": "Social security charged per attendance day", "verify_qty": True, "tolerance": 0.5},
    "name_matching": {"threshold": 0.35, "auto_approve_unmatched_hours": 0.5, "flag_unmatched": True}
  },
  "item_classification": {
    "work_hours_items": ["Prestations", "Hrs suppl\u00e9m. pay\u00e9es imm\u00e9diatement"],
    "supplement_items": ["Suppl\u00e9ment \u00e9quipe"],
    "social_items": ["Dimona"],
    "ignore_items": ["Prime de pension", "Frais domicile-travail", "Ch\u00e8ques repas", "Eco-ch\u00e8ques", "Surcharge frais admin. \u00e9tudiant", "Frais de gestion \u00e9tudiant"]
  }
}
with open(r'project/rules/belgium/tempoteam.json', 'w', encoding='utf-8') as f:
    json.dump(rules, f, indent=2, ensure_ascii=False)
print("Rules written")

# Write server.py
server_py = '''"""Flask web server for Invoice-Attendance Reconciliation."""
import os, sys, json, re, tempfile, shutil, uuid
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "app"))
from flask import Flask, request, jsonify, render_template, session, send_file
from werkzeug.utils import secure_filename
from parsers.attendance_parser import parse_attendance
from parsers.invoice_parser import parse_invoices
from reconciliation import reconcile, _score_name
import io, base64

app = Flask(__name__)
app.secret_key = os.urandom(24)
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
RULES_DIR = os.path.join(os.path.dirname(__file__), "rules")

# -------- HELPERS --------
def get_available_suppliers():
    suppliers = []
    for root, dirs, files in os.walk(RULES_DIR):
        for f in files:
            if f.endswith(".json"):
                path = os.path.join(root, f)
                try:
                    with open(path, "r", encoding="utf-8") as fp:
                        data = json.load(fp)
                    suppliers.append({"country": data.get("country", ""), "supplier": data.get("supplier", ""), "label": data.get("label", ""), "file": os.path.relpath(path, RULES_DIR)})
                except: pass
    return suppliers

def load_rule(country, supplier):
    for root, dirs, files in os.walk(RULES_DIR):
        for f in files:
            if f.endswith(".json"):
                path = os.path.join(root, f)
                try:
                    with open(path, "r", encoding="utf-8") as fp:
                        data = json.load(fp)
                    if data.get("country", "").lower() == country.lower() and data.get("supplier", "").lower() == supplier.lower():
                        return data
                except: pass
    return None

# -------- ROUTES --------
@app.route("/")
def index():
    suppliers = get_available_suppliers()
    return render_template("index.html", suppliers=suppliers)

@app.route("/api/suppliers")
def api_suppliers():
    return jsonify(get_available_suppliers())

@app.route("/api/rules/<country>/<supplier>")
def api_get_rule(country, supplier):
    rule = load_rule(country, supplier)
    if rule:
        return jsonify(rule)
    return jsonify({"error": "Rule not found"}), 404

@app.route("/api/rules/<country>/<supplier>", methods=["PUT"])
def api_update_rule(country, supplier):
    data = request.get_json()
    rule = load_rule(country, supplier)
    if not rule:
        return jsonify({"error": "Rule not found"}), 404
    # Find file
    for root, dirs, files in os.walk(RULES_DIR):
        for f in files:
            if f.endswith(".json"):
                path = os.path.join(root, f)
                try:
                    with open(path, "r", encoding="utf-8") as fp:
                        existing = json.load(fp)
                    if existing.get("country","").lower() == country.lower() and existing.get("supplier","").lower() == supplier.lower():
                        with open(path, "w", encoding="utf-8") as fp:
                            json.dump(data, fp, indent=2, ensure_ascii=False)
                        return jsonify({"status": "ok"})
                except: pass
    return jsonify({"error": "File not found"}), 404

@app.route("/api/rules", methods=["POST"])
def api_create_rule():
    data = request.get_json()
    country = data.get("country", "unknown").lower()
    supplier = data.get("supplier", "unknown").upper().replace(" ", "_")
    dir_path = os.path.join(RULES_DIR, country)
    os.makedirs(dir_path, exist_ok=True)
    filepath = os.path.join(dir_path, supplier.lower() + ".json")
    data["country"] = country
    data["supplier"] = supplier
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    return jsonify({"status": "created", "file": os.path.relpath(filepath, RULES_DIR)})

@app.route("/api/rules/countries")
def api_list_countries():
    countries = set()
    for root, dirs, files in os.walk(RULES_DIR):
        for f in files:
            if f.endswith(".json"):
                try:
                    with open(os.path.join(root, f), "r", encoding="utf-8") as fp:
                        data = json.load(fp)
                    countries.add((data.get("country", ""), data.get("label", "")))
                except: pass
    result = [{"country": c, "label": l} for c, l in sorted(countries)]
    return jsonify(result)

@app.route("/api/parse-attendance", methods=["POST"])
def api_parse_attendance():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    country = request.form.get("country", "belgium")
    supplier = request.form.get("supplier", "TEMPOTEAM")
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    f.save(tmp.name)
    try:
        att = parse_attendance(tmp.name, country=country, supplier=supplier)
        records = []
        for rec in att.records:
            records.append({"name": rec.employee_name, "date": rec.date, "hours": rec.hours, "status": rec.status, "raw_time": rec.raw_time_slot, "role": rec.role})
        result = {"period_start": att.period_start, "period_end": att.period_end, "total_hours": round(att.get_total_hours(), 2), "employee_count": len(att.get_employees()), "employees": att.get_employees(), "records": records}
        session["attendance_data"] = result
        session["attendance_file"] = f.filename
        os.unlink(tmp.name)
        return jsonify(result)
    except Exception as e:
        os.unlink(tmp.name)
        return jsonify({"error": str(e)}), 500

@app.route("/api/parse-invoices", methods=["POST"])
def api_parse_invoices():
    if "files" not in request.files:
        return jsonify({"error": "No files"}), 400
    files = request.files.getlist("files")
    tmp_paths = []
    try:
        inv_paths = []
        for f in files:
            tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            f.save(tmp.name)
            tmp_paths.append(tmp.name)
            inv_paths.append(tmp.name)
        invs = parse_invoices(inv_paths)
        result = {}
        for num, inv in invs.items():
            emps = []
            for e in inv.emps:
                items = [{"name": i.name, "qty": round(i.qty, 2), "rate": round(i.rate, 4), "pct": round(i.pct, 2), "amt": round(i.amt, 2)} for i in e.items]
                emps.append({"name": e.name, "hours": round(e.hours(), 2), "role": e.role, "period": e.period, "items": items, "subtotal": round(e.subtotal, 2)})
            result[num] = {"number": num, "date": inv.date, "period_start": inv.p_start, "period_end": inv.p_end, "type": inv.type, "total_hours": round(inv.total_hours(), 2), "excl": round(inv.excl, 2), "incl": round(inv.incl, 2), "employees": emps}
        session["invoice_data"] = result
        for p in tmp_paths:
            try: os.unlink(p)
            except: pass
        return jsonify(result)
    except Exception as e:
        for p in tmp_paths:
            try: os.unlink(p)
            except: pass
        return jsonify({"error": str(e)}), 500

@app.route("/api/reconcile", methods=["POST"])
def api_reconcile():
    data = request.get_json() or {}
    country = data.get("country", "belgium")
    supplier = data.get("supplier", "TEMPOTEAM")
    att_data = session.get("attendance_data")
    inv_data = session.get("invoice_data")
    if not att_data or not inv_data:
        return jsonify({"error": "Missing attendance or invoice data. Run parsing first."}), 400

    # Load rule
    rule = load_rule(country, supplier)
    if not rule:
        rule = {"rules": {"name_matching": {"threshold": 0.35}}}

    # Re-parse for actual reconcile
    import tempfile
    import openpyxl
    tmp_xlsx = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Juin 2026"
        ws.cell(2, 2).value = att_data["period_start"]
        ws.cell(2, 8).value = att_data["period_end"]
        seen_emps = {}
        emp_row = 3
        for rec in att_data["records"]:
            name = rec["name"]
            if name not in seen_emps:
                seen_emps[name] = emp_row
                ws.cell(emp_row, 1).value = name
                emp_row += 1
        day_map = {15:2,16:3,17:4,18:5,19:6,20:7,21:8,22:2,23:3,24:4,25:5,26:6,27:7,28:8}
        for rec in att_data["records"]:
            name = rec["name"]
            row = seen_emps.get(name)
            if not row: continue
            d = int(rec["date"].split("-")[2])
            col = day_map.get(d)
            if col and rec["raw_time"]:
                ws.cell(row, col).value = rec["raw_time"]
        wb.save(tmp_xlsx)
        att = parse_attendance(tmp_xlsx)
        os.unlink(tmp_xlsx)
    except Exception as e:
        try: os.unlink(tmp_xlsx)
        except: pass
        return jsonify({"error": "Failed to rebuild attendance: " + str(e)}), 500

    # Re-parse invoices
    tmp_pdfs = []
    try:
        for inv_num, inv_info in inv_data.items():
            tmp_pdf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            # Create a minimal PDF with invoice text for re-parsing
            from reportlab.pdfgen import canvas
            c = canvas.Canvas(tmp_pdf.name)
            c.drawString(50, 800, "Facture " + inv_num)
            c.drawString(50, 780, "Numero de facture " + inv_num)
            c.drawString(50, 760, "Periode " + att_data["period_start"] + " - " + att_data["period_end"])
            for emp in inv_info.get("employees", []):
                y = 720
                c.drawString(50, y, emp["name"])
                y -= 20
                for item in emp.get("items", []):
                    c.drawString(50, y, "{} {} {} {} {}".format(item["name"], item["qty"], item["rate"], item["pct"], item["amt"]))
                    y -= 15
            c.save()
            tmp_pdfs.append(tmp_pdf.name)
        invs = parse_invoices(tmp_pdfs)
    except Exception as e:
        for p in tmp_pdfs:
            try: os.unlink(p)
            except: pass
        return jsonify({"error": "Failed to re-parse invoices: " + str(e)}), 500

    try:
        report = reconcile(att, list(invs.values()))
        results = []
        for r in report.results:
            results.append({
                "name": r["name"], "att_name": r["att_name"],
                "att_hours": r["att_hours"], "att_days": r.get("att_days", 0),
                "inv_hours": r["inv_hours"], "inv_amount": r["inv_amount"],
                "diff_hours": r["diff_hours"], "diff_percent": r["diff_percent"],
                "verdict": r["verdict"],
                "supplement_check": r.get("supplement_check"),
                "dimona_check": r.get("dimona_check"),
                "items": r.get("items_breakdown", []),
                "unmatched": r.get("unmatched", False),
            })
        summary = {
            "period": report.period, "supplier": report.supplier,
            "total_att": round(report.total_attendance_hours, 2),
            "total_inv": round(report.total_invoice_hours, 2),
            "total_amt": round(report.total_invoice_amount, 2),
            "employees": len(report.results),
            "results": results,
            "unmatched_att": report.unmatched_attendance,
            "unmatched_inv": report.unmatched_invoice,
        }
        for p in tmp_pdfs:
            try: os.unlink(p)
            except: pass
        return jsonify(summary)
    except Exception as e:
        for p in tmp_pdfs:
            try: os.unlink(p)
            except: pass
        raise

if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
'''

with open(r'project/server.py', 'w', encoding='utf-8') as f:
    f.write(server_py)
print("server.py written")
