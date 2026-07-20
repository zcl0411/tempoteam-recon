"""Flask web server for Invoice-Attendance Reconciliation."""
import os, sys, json, re, tempfile, shutil, uuid
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "app"))
from flask import Flask, request, jsonify, render_template, session, send_file
from werkzeug.utils import secure_filename
from parsers.attendance_parser import parse_attendance
from parsers.invoice_parser import parse_invoices
from reconciliation import reconcile, _score_name
import io, base64

app = Flask(__name__)
app.secret_key = os.urandom(24)
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
RULES_DIR = os.path.join(os.path.dirname(__file__), "rules")

# -------- HELPERS --------
def get_available_suppliers():
    suppliers = []
    for root, dirs, files in os.walk(RULES_DIR):
        for f in files:
            if f.endswith(".json"):
                path = os.path.join(root, f)
                try:
                    with open(path, "r", encoding="utf-8") as fp:
                        data = json.load(fp)
                    suppliers.append({"country": data.get("country", ""), "supplier": data.get("supplier", ""), "label": data.get("label", ""), "file": os.path.relpath(path, RULES_DIR)})
                except: pass
    return suppliers

def load_rule(country, supplier):
    for root, dirs, files in os.walk(RULES_DIR):
        for f in files:
            if f.endswith(".json"):
                path = os.path.join(root, f)
                try:
                    with open(path, "r", encoding="utf-8") as fp:
                        data = json.load(fp)
                    if data.get("country", "").lower() == country.lower() and data.get("supplier", "").lower() == supplier.lower():
                        return data
                except: pass
    return None

# -------- ROUTES --------
@app.route("/")
def index():
    suppliers = get_available_suppliers()
    return render_template("index.html", suppliers=suppliers)

@app.route("/api/suppliers")
def api_suppliers():
    return jsonify(get_available_suppliers())

@app.route("/api/rules/<country>/<supplier>")
def api_get_rule(country, supplier):
    rule = load_rule(country, supplier)
    if rule:
        return jsonify(rule)
    return jsonify({"error": "Rule not found"}), 404

@app.route("/api/rules/<country>/<supplier>", methods=["PUT"])
def api_update_rule(country, supplier):
    data = request.get_json()
    rule = load_rule(country, supplier)
    if not rule:
        return jsonify({"error": "Rule not found"}), 404
    # Find file
    for root, dirs, files in os.walk(RULES_DIR):
        for f in files:
            if f.endswith(".json"):
                path = os.path.join(root, f)
                try:
                    with open(path, "r", encoding="utf-8") as fp:
                        existing = json.load(fp)
                    if existing.get("country","").lower() == country.lower() and existing.get("supplier","").lower() == supplier.lower():
                        with open(path, "w", encoding="utf-8") as fp:
                            json.dump(data, fp, indent=2, ensure_ascii=False)
                        return jsonify({"status": "ok"})
                except: pass
    return jsonify({"error": "File not found"}), 404

@app.route("/api/rules", methods=["POST"])
def api_create_rule():
    data = request.get_json()
    country = data.get("country", "unknown").lower()
    supplier = data.get("supplier", "unknown").upper().replace(" ", "_")
    dir_path = os.path.join(RULES_DIR, country)
    os.makedirs(dir_path, exist_ok=True)
    filepath = os.path.join(dir_path, supplier.lower() + ".json")
    data["country"] = country
    data["supplier"] = supplier
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    return jsonify({"status": "created", "file": os.path.relpath(filepath, RULES_DIR)})

@app.route("/api/rules/countries")
def api_list_countries():
    countries = set()
    for root, dirs, files in os.walk(RULES_DIR):
        for f in files:
            if f.endswith(".json"):
                try:
                    with open(os.path.join(root, f), "r", encoding="utf-8") as fp:
                        data = json.load(fp)
                    countries.add((data.get("country", ""), data.get("label", "")))
                except: pass
    result = [{"country": c, "label": l} for c, l in sorted(countries)]
    return jsonify(result)

@app.route("/api/parse-attendance", methods=["POST"])
def api_parse_attendance():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    country = request.form.get("country", "belgium")
    supplier = request.form.get("supplier", "TEMPOTEAM")
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    f.save(tmp.name)
    try:
        att = parse_attendance(tmp.name, country=country, supplier=supplier)
        records = []
        for rec in att.records:
            records.append({"name": rec.employee_name, "date": rec.date, "hours": rec.hours, "status": rec.status, "raw_time": rec.raw_time_slot, "role": rec.role})
        result = {"period_start": att.period_start, "period_end": att.period_end, "total_hours": round(att.get_total_hours(), 2), "employee_count": len(att.get_employees()), "employees": att.get_employees(), "records": records}
        import uuid
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        sid = session.get("session_id") or str(uuid.uuid4())
        session["session_id"] = sid
        att_path = os.path.join(UPLOAD_FOLDER, sid + "_attendance.json")
        with open(att_path, "w", encoding="utf-8") as _f:
            json.dump(result, _f, ensure_ascii=False)
        tmp.close()  # Release file handle on Windows
        os.unlink(tmp.name)
        return jsonify(result)
    except Exception as e:
        tmp.close()  # Release file handle on Windows
        os.unlink(tmp.name)
        return jsonify({"error": str(e)}), 500

@app.route("/api/parse-invoices", methods=["POST"])
def api_parse_invoices():
    if "files" not in request.files:
        return jsonify({"error": "No files"}), 400
    files = request.files.getlist("files")
    tmp_paths = []
    try:
        inv_paths = []
        for f in files:
            tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            f.save(tmp.name)
            tmp.close()  # Release handle on Windows
            tmp_paths.append(tmp.name)
            inv_paths.append(tmp.name)
        invs = parse_invoices(inv_paths)
        result = {}
        for num, inv in invs.items():
            emps = []
            for e in inv.emps:
                items = [{"name": i.name, "qty": round(i.qty, 2), "rate": round(i.rate, 4), "pct": round(i.pct, 2), "amt": round(i.amt, 2)} for i in e.items]
                emps.append({"name": e.name, "hours": round(e.hours(), 2), "role": e.role, "period": e.period, "items": items, "subtotal": round(e.subtotal, 2)})
            result[num] = {"number": num, "date": inv.date, "period_start": inv.p_start, "period_end": inv.p_end, "type": inv.type, "total_hours": round(inv.total_hours(), 2), "excl": round(inv.excl, 2), "incl": round(inv.incl, 2), "employees": emps}
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        sid = session.get("session_id")
        if sid:
            inv_path = os.path.join(UPLOAD_FOLDER, sid + "_invoice.json")
            with open(inv_path, "w", encoding="utf-8") as _f:
                json.dump(result, _f, ensure_ascii=False)
        for p in tmp_paths:
            try: os.unlink(p)
            except: pass
        return jsonify(result)
    except Exception as e:
        for p in tmp_paths:
            try: os.unlink(p)
            except: pass
        return jsonify({"error": str(e)}), 500

@app.route("/api/reconcile", methods=["POST"])
def api_reconcile():
    data = request.get_json() or {}
    country = data.get("country", "belgium")
    supplier = data.get("supplier", "TEMPOTEAM")
    sid = session.get("session_id")
    att_data = None
    inv_data = None
    if sid:
        att_path = os.path.join(UPLOAD_FOLDER, sid + "_attendance.json")
        inv_path = os.path.join(UPLOAD_FOLDER, sid + "_invoice.json")
        if os.path.exists(att_path):
            with open(att_path, "r", encoding="utf-8") as _f:
                att_data = json.load(_f)
        if os.path.exists(inv_path):
            with open(inv_path, "r", encoding="utf-8") as _f:
                inv_data = json.load(_f)
    if not att_data or not inv_data:
        return jsonify({"error": "Missing attendance or invoice data. Run parsing first."}), 400

    # Load rule
    rule = load_rule(country, supplier)
    if not rule:
        rule = {"rules": {"name_matching": {"threshold": 0.35}}}

    # Re-parse for actual reconcile
    import tempfile
    import openpyxl
    tmp_xlsx = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Juin 2026"
        ws.cell(2, 2).value = att_data["period_start"]
        ws.cell(2, 8).value = att_data["period_end"]
        seen_emps = {}
        emp_row = 3
        for rec in att_data["records"]:
            name = rec["name"]
            if name not in seen_emps:
                seen_emps[name] = emp_row
                ws.cell(emp_row, 1).value = name
                emp_row += 1
        from datetime import datetime
        day_map = {}
        for rec in att_data["records"]:
            d = int(rec["date"].split("-")[2])
            if d in day_map: continue
            dt = datetime.strptime(rec["date"], "%Y-%m-%d")
            col = dt.weekday() + 2  # Mon=0->2, Tue=1->3, ..., Sun=6->8
            day_map[d] = col
            ws.cell(2, col).value = d  # Set date number in header row
        for rec in att_data["records"]:
            name = rec["name"]
            row = seen_emps.get(name)
            if not row: continue
            d = int(rec["date"].split("-")[2])
            col = day_map.get(d)
            if col and rec["raw_time"]:
                ws.cell(row, col).value = rec["raw_time"]
        wb.save(tmp_xlsx)
        att = parse_attendance(tmp_xlsx.name)
        tmp_xlsx.close()  # Release handle on Windows
        os.unlink(tmp_xlsx.name)
    except Exception as e:
        try: tmp_xlsx.close(); os.unlink(tmp_xlsx.name)
        except: pass
        return jsonify({"error": "Failed to rebuild attendance: " + str(e)}), 500

    # Use already-parsed invoice data directly (skip re-parsing)
    from parsers.invoice_parser import Invoice, EmpDetail, InvItem
    invs = {}
    for inv_num, inv_info in inv_data.items():
        inv = Invoice()
        inv.num = inv_info.get("number", inv_num)
        inv.date = inv_info.get("date", "")
        inv.p_start = inv_info.get("period_start", "")
        inv.p_end = inv_info.get("period_end", "")
        inv.excl = inv_info.get("excl", 0)
        inv.incl = inv_info.get("incl", 0)
        inv.type = inv_info.get("type", "")
        for ed in inv_info.get("employees", []):
            emp = EmpDetail(name=ed.get("name", ""), period=ed.get("period", ""), role=ed.get("role", ""))
            emp.subtotal = ed.get("subtotal", 0)
            for item in ed.get("items", []):
                emp.add(InvItem(
                    name=item.get("name", ""),
                    qty=item.get("qty", 0),
                    rate=item.get("rate", 0),
                    pct=item.get("pct", 0),
                    amt=item.get("amt", 0),
                    vat=item.get("vat", "A")
                ))
            if emp.items or emp.name:
                inv.add(emp)
        if inv.num:
            invs[inv.num] = inv

    try:
        report = reconcile(att, list(invs.values()))
        results = []
        for r in report.results:
            results.append({
                "name": r["name"], "att_name": r["att_name"],
                "att_hours": r["att_hours"], "att_days": r.get("att_days", 0),
                "inv_hours": r["inv_hours"], "inv_amount": r["inv_amount"],
                "diff_hours": r["diff_hours"], "diff_percent": r["diff_percent"],
                "verdict": r["verdict"],
                "supplement_check": r.get("supplement_check"),
                "dimona_check": r.get("dimona_check"),
                "items": r.get("items_breakdown", []),
                "unmatched": r.get("unmatched", False),
            })
        summary = {
            "period": report.period, "supplier": report.supplier,
            "total_att": round(report.total_attendance_hours, 2),
            "total_inv": round(report.total_invoice_hours, 2),
            "total_amt": round(report.total_invoice_amount, 2),
            "employees": len(report.results),
            "results": results,
            "unmatched_att": report.unmatched_attendance,
            "unmatched_inv": report.unmatched_invoice,
        }
        sid = session.get("session_id")
        if sid:
            res_path = os.path.join(UPLOAD_FOLDER, sid + "_results.json")
            with open(res_path, "w", encoding="utf-8") as _f:
                json.dump(summary, _f, ensure_ascii=False)
        return jsonify(summary)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/export-review", methods=["POST"])
def api_export_review():
    """Export manual review data as XLSX."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    sid = session.get("session_id")
    if not sid:
        return jsonify({"error": "No session data"}), 400
    
    res_path = os.path.join(UPLOAD_FOLDER, sid + "_results.json")
    if not os.path.exists(res_path):
        return jsonify({"error": "No results. Run reconciliation first."}), 400
    
    with open(res_path, "r", encoding="utf-8") as _f:
        summary = json.load(_f)
    
    review_items = [r for r in summary.get("results", [])
                    if r.get("verdict") in ("manual_review", "mismatch")]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人工复核"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A1F35", end_color="1A1F35", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="2A3050"),
        right=Side(style="thin", color="2A3050"),
        top=Side(style="thin", color="2A3050"),
        bottom=Side(style="thin", color="2A3050"),
    )
    headers = ["员工姓名", "考勤工时", "发票工时", "差异", "判定", "班次补贴", "补贴工时(考勤)", "补贴工时(发票)", "社保Dimona", "发票明细"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    
    for i, item in enumerate(review_items, 2):
        supp = item.get("supplement_check") or {}
        dimona = item.get("dimona_check") or {}
        
        supp_map = {"ok": "正常", "missing": "缺失", "amount_mismatch": "金额不符", "unexpected": "异常"}
        dimona_map = {"ok": "正常", "qty_mismatch": "数量不符"}
        verdict_map = {"auto_approved": "自动通过", "match": "一致", "minor_diff": "轻微差异", "manual_review": "查看数据", "mismatch": "异常"}
        
        supp_status = supp_map.get(supp.get("status"), supp.get("status", ""))
        dimona_status = dimona_map.get(dimona.get("status"), dimona.get("status", ""))
        verdict_label = verdict_map.get(item.get("verdict", ""), item.get("verdict", ""))
        
        supp_extra = ""
        if supp.get("status") and supp["status"] not in ("ok", None):
            supp_extra = " (考勤=%.2fh, 发票=%.2fh)" % (supp.get("att_hours", 0), supp.get("inv_hours", 0))
        dimona_extra = ""
        if dimona.get("status") and dimona["status"] != "ok":
            dimona_extra = " (发票qty=%s, 预期qty=%s)" % (dimona.get("inv_qty", "?"), dimona.get("expected_qty", "?"))
        
        items_str = "; ".join(item.get("items", []))
        
        ws.cell(i, 1, item.get("att_name", ""))
        ws.cell(i, 2, item.get("att_hours", 0))
        ws.cell(i, 3, item.get("inv_hours", 0))
        ws.cell(i, 4, round(item.get("diff_hours", 0), 2))
        ws.cell(i, 5, verdict_label)
        ws.cell(i, 6, supp_status + supp_extra)
        ws.cell(i, 7, supp.get("att_hours", "-"))
        ws.cell(i, 8, supp.get("inv_hours", "-"))
        ws.cell(i, 9, dimona_status + dimona_extra)
        ws.cell(i, 10, items_str)
        
        for col in range(1, 11):
            cell = ws.cell(i, col)
            cell.border = thin_border
            cell.alignment = Alignment(wrapText=True)
    
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 35
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 35
    ws.column_dimensions["J"].width = 50
    
    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    
    with open(tmp.name, "rb") as _f:
        data = _f.read()
    os.unlink(tmp.name)
    
    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="人工复核数据.xlsx",
    )

if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)

