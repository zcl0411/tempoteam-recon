"""Attendance Sheet Parser
Supports French/Belgian time format with options for lunch break deduction.
"""
import re
import openpyxl
from typing import List, Dict, Optional

class AttendanceRecord:
    def __init__(self, employee_name: str, date: str, hours: float, 
                 role: str = "", status: str = "present", raw_time_slot: str = ""):
        self.employee_name = employee_name
        self.date = date
        self.hours = hours
        self.role = role
        self.status = status
        self.raw_time_slot = raw_time_slot

class AttendanceSheet:
    def __init__(self, period_start: str = "", period_end: str = ""):
        self.period_start = period_start
        self.period_end = period_end
        self.records: List[AttendanceRecord] = []
    def add_record(self, record):
        self.records.append(record)
    def get_hours_by_employee(self, name: str) -> float:
        return sum(r.hours for r in self.records 
                   if r.employee_name.lower().strip() == name.lower().strip()
                   and r.status == "present")
    def get_total_hours(self) -> float:
        return sum(r.hours for r in self.records if r.status == "present")
    def get_employees(self) -> List[str]:
        seen = set(); result = []
        for r in self.records:
            n = r.employee_name.strip()
            if n and n not in seen: seen.add(n); result.append(n)
        return result
    def get_records_by_employee(self, name: str) -> List[AttendanceRecord]:
        return [r for r in self.records if r.employee_name.lower().strip() == name.lower().strip()]
    def get_attendance_days(self, name: str) -> int:
        return sum(1 for r in self.records
                   if r.employee_name.lower().strip() == name.lower().strip()
                   and r.status == "present")

TIME_PATTERN = re.compile(r"(\d+)h(\d*)\s*[-\u2013]\s*(\d+)h(\d*)")

def detect_shift_start(time_str: str) -> int:
    """Extract the start hour from a French time slot like '6h-14h06' -> returns 6."""
    m = re.match(r"(\d+)h", time_str.strip())
    return int(m.group(1)) if m else -1

def parse_french_time_slot(time_str: str, deduct_lunch: bool = True) -> float:
    if not time_str or not isinstance(time_str, str):
        return 0.0
    time_str = time_str.strip()
    upper = time_str.upper()
    if upper in ("OFF", "ABSENT", "CONGÉ", "FÉRIÉ", "MALADE"):
        return 0.0
    match = TIME_PATTERN.search(time_str)
    if not match:
        return 0.0
    start_h = int(match.group(1))
    start_m = int(match.group(2)) if match.group(2) else 0
    end_h = int(match.group(3))
    end_m = int(match.group(4)) if match.group(4) else 0
    if end_h < start_h or (end_h == start_h and end_m < start_m):
        end_h += 24
    total_minutes = (end_h * 60 + end_m) - (start_h * 60 + start_m)
    if deduct_lunch and total_minutes >= 360:
        total_minutes -= 30
    return round(total_minutes / 60, 2)

def extract_role_from_name(name: str) -> str:
    match = re.search(r"\(([^)]+)\)", name)
    if match:
        role_text = match.group(1)
        role_map = {
            "Conducteur de chariot élévateur": ["Cariste", "chariot élévateur", "CARISTE"],
            "Douane": ["Douane", "douane"],
            "Manoeuvre": ["Manoeuvre", "WH", "Stack", "STACK", "stack"],
            "Étudiant": ["TT", "étudiante", "étudiant", "Retour", "dispatch"],
        }
        for role_name, keywords in role_map.items():
            for kw in keywords:
                if kw.lower() in role_text.lower():
                    return role_name
        return role_text
    return ""

def parse_tempoteam_attendance(filepath: str, config: dict = None) -> AttendanceSheet:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_name = "Juin 2026"
    name_col = 1
    day_cols = {2: "Mon", 3: "Tue", 4: "Wed", 5: "Thu", 6: "Fri", 7: "Sat", 8: "Sun"}
    
    ws = wb[sheet_name]
    date_numbers = {}
    for c in range(2, 9):
        v = ws.cell(2, c).value
        if v: date_numbers[c] = int(v)
    
    sheet = AttendanceSheet()
    if ws.cell(2, 2).value and ws.cell(2, 8).value:
        sheet.period_start = str(ws.cell(2, 2).value)
        sheet.period_end = str(ws.cell(2, 8).value)
    
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        role = extract_role_from_name(name)
        
        for col, day_name in day_cols.items():
            time_val = ws.cell(r, col).value
            if not time_val:
                continue
            time_str = str(time_val).strip()
            
            # Check status first
            upper = time_str.upper()
            if upper == "OFF":
                status = "off"
                hours = 0.0
            elif upper == "ABSENT":
                status = "absent"
                hours = 0.0
            else:
                status = "present"
                hours = parse_french_time_slot(time_str, deduct_lunch=True)
            raw_time_slot = time_str
            
            month = 6
            day = date_numbers.get(col, 0)
            if day > 0:
                date_str = f"2026-{month:02d}-{day:02d}"
                sheet.add_record(AttendanceRecord(name, date_str, hours, role, status, raw_time_slot))
    
    return sheet

def parse_attendance(filepath: str, country: str = "belgium", 
                     supplier: str = "TEMPOTEAM", config: dict = None) -> AttendanceSheet:
    return parse_tempoteam_attendance(filepath, config)

if __name__ == "__main__":
    import os
    d1 = r"C:\Users\zt26501\Documents\Codex\2026-07-17\pdf-pdf-pdf-pdf\inputs\period1"
    s = parse_attendance(os.path.join(d1, "0615-0621.xlsx"))
    print(f"P1: {s.get_total_hours():.2f}h ({len(s.get_employees())} emps)")
    for e in s.get_employees()[:5]:
        print(f"  {e}: {s.get_hours_by_employee(e):.2f}h")
    
    d2 = r"C:\Users\zt26501\Documents\Codex\2026-07-17\pdf-pdf-pdf-pdf\inputs\period2"
    s2 = parse_attendance(os.path.join(d2, "0622-0628.xlsx"))
    print(f"\nP2: {s2.get_total_hours():.2f}h ({len(s2.get_employees())} emps)")
    for e in s2.get_employees()[:5]:
        print(f"  {e}: {s2.get_hours_by_employee(e):.2f}h")